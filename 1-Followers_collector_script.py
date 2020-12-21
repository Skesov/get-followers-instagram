from dotenv import load_dotenv
import instaloader
import openpyxl
import os
import time

load_dotenv()

# Files saving settings

path_to_save = 'Output'
followers_filename = 'Followers.xlsx'

# Get instance

user = os.getenv('LOGIN')
profile = os.getenv('PROFILE_FOR_PARSING')

L = instaloader.Instaloader()

# Load session

L.load_session_from_file(user)
# (load session created w/
#  `instaloader -l USERNAME`)

# Obtain profile metadata
profile = instaloader.Profile.from_username(L.context, profile)


# Checking folders existing

checkDirEx = os.path.isdir(path_to_save)
if not checkDirEx:
    os.mkdir(path_to_save)

# # Create excel sheet

book = openpyxl.Workbook()
sheet = book.active
sheet = book.create_sheet('Followers', 0)
sheet.column_dimensions['A'].width = 10
sheet.column_dimensions['B'].width = 25
sheet.column_dimensions['C'].width = 30
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 10
sheet.column_dimensions['G'].width = 20
sheet.column_dimensions['H'].width = 10
sheet.column_dimensions['I'].width = 15


# Creating a first row with titles
sheet.cell(1, 1).value = '№'
sheet.cell(1, 2).value = 'Username'
sheet.cell(1, 3).value = 'Fullname'
sheet.cell(1, 4).value = 'Posts'
sheet.cell(1, 5).value = 'Followers'
sheet.cell(1, 6).value = 'Followees'
sheet.cell(1, 7).value = 'Followee/Folowers'
sheet.cell(1, 8).value = 'I follow'
sheet.cell(1, 9).value = 'ID'


# collecting folowees
print(f'Collecting followees')
list_of_followees = []

for followee in profile.get_followees():
    list_of_followees.append(followee.username)


print(f'Followees collected')

# collecting folowers and writing information to file
row = 2
print(f'collecting followers')
for follower in profile.get_followers():
    print(f'Working with follower #{row-1}')
    username = follower.username
    ifollow = 'yes' if username in list_of_followees else 'no'

    sheet.cell(row, 1).value = row-1
    sheet.cell(row, 2).value = username
    sheet.cell(row, 8).value = ifollow

    row += 1

print(f'Followers collected')


# saving book
print(f'Saving {followers_filename}.xlsx in {path_to_save}')
book.save(path_to_save + "/" + followers_filename)

print("\n __________")
print("|          |")
print("|   DONE   |")
print("|__________|")
