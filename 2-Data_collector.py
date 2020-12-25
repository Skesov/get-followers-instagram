import os
import time
from random import randint

import instaloader
import openpyxl
from dotenv import load_dotenv

load_dotenv()

# Files saving settings

path = 'Output'
followers_filename = 'Followers.xlsx'
fullpath = path + '/' + followers_filename

# Get instance

user = os.getenv('LOGIN')
profile = os.getenv('PROFILE_FOR_PARSING')

L = instaloader.Instaloader()

# Load session

L.load_session_from_file(user)
# (load session created w/
#  `instaloader -l USERNAME`)

# Checking file existing

if os.path.isfile(fullpath):
    print('File exist')
else:
    print('File not exist or have another name')
    exit('\nNot have a file for execuation')

# Open excel sheet


book = openpyxl.load_workbook(fullpath)
print(f'Book opened')
sheet = book.active
row_count = sheet.max_row
print(f'{row_count-1} followers')

row = 2

while row <= row_count:
    username = sheet.cell(row, 2).value
    print(f'Read information about {username} in line {row-1}')
    if sheet.cell(row, 5).value == None:
        try:
            profile = instaloader.Profile.from_username(L.context, username)
            fullname = profile.full_name
            number_of_posts = profile.mediacount
            number_of_followers = profile.followers
            number_of_followees = profile.followees
            profileID = profile.userid

            if number_of_followers != 0:
                ratio = number_of_followees / number_of_followers
            else:
                ratio = 'Bot'

            sheet.cell(row, 3).value = fullname
            sheet.cell(row, 4).value = number_of_posts
            sheet.cell(row, 5).value = number_of_followers
            sheet.cell(row, 6).value = number_of_followees
            sheet.cell(row, 7).value = ratio
            sheet.cell(row, 9).value = profileID

            # saving book
            timer = randint(15, 35)
            print(f'Saving book. Go to next line in {timer} seconds')
            book.save(fullpath)
            time.sleep(timer)
            row += 1
        except Exception as error:
            print('Something went wrong:')
            print(error)
            print(type(error))
            if isinstance(error, instaloader.exceptions.ProfileNotExistsException):
                print('User not exist or banned. Next line.')
                time.sleep(20)
                row += 1
            elif isinstance(error, instaloader.exceptions.LoginRequiredException):
                print('Redirect to login page. Next line.')
                time.sleep(20)
                row += 1
            elif isinstance(error, instaloader.exceptions.ConnectionException):
                sleep_timer = randint(2700, 4000)
                current_time = time.mktime(time.localtime())
                added_time = time.localtime(current_time + sleep_timer)
                time_retring = time.strftime("%H:%M:%S", added_time)

                print(
                    f'Connection exception Waiting {int(sleep_timer / 60)} minutes.')
                print(f'The request will be retried at {time_retring}')
                time.sleep(sleep_timer)

            elif isinstance(error, instaloader.exceptions.QueryReturnedBadRequestException):
                print('Blocked account. Need actions on instagram page.')
                retry = input('Input anything for retry')

            else:
                sleep_timer = randint(1500, 2000)
                current_time = time.mktime(time.localtime())
                added_time = time.localtime(current_time + sleep_timer)
                time_retring = time.strftime("%H:%M:%S", added_time)

                print(f'\nWaiting {time_retring} minutes')
                time.sleep(sleep_timer)

    else:
        print('Data is exist. Next line')
        row += 1


print("\n__________")
print("|          |")
print("|   DONE   |")
print("|__________|")
